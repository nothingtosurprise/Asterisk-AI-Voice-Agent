import csv
import io
import json

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook
import pytest

from api import outbound as outbound_api


class _FakeOutboundStore:
    def __init__(self):
        self.imports = []

    async def import_leads_csv(self, campaign_id, csv_bytes, **kwargs):
        self.imports.append((campaign_id, csv_bytes, kwargs))
        return {
            "accepted": 1,
            "rejected": 0,
            "duplicates": 0,
            "errors": [],
            "error_csv": "",
            "error_csv_truncated": False,
            "warnings": [],
            "warnings_truncated": False,
        }


def _xlsx_bytes(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leads"
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _client(monkeypatch):
    store = _FakeOutboundStore()
    monkeypatch.setattr(outbound_api, "_get_outbound_store", lambda: store)
    monkeypatch.setattr(
        outbound_api,
        "_load_known_agent_selectors",
        lambda: (["sales"], ["Sales Team", "sales"]),
    )
    app = FastAPI()
    app.include_router(outbound_api.router, prefix="/api")
    return TestClient(app), store


def test_xlsx_conversion_uses_first_sheet_and_preserves_zero_number_format():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["name", "phone_number", "agent"])
    sheet.append(["Extension", 2765, "sales"])
    sheet["B2"].number_format = "00000"
    workbook.create_sheet("Ignored").append(["phone_number", "9999"])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    converted = outbound_api._xlsx_to_csv_bytes(output.getvalue()).decode("utf-8")
    rows = list(csv.reader(io.StringIO(converted)))

    assert rows == [
        ["name", "phone_number", "agent"],
        ["Extension", "02765", "sales"],
    ]


def test_xlsx_import_endpoint_normalizes_into_existing_csv_path(monkeypatch):
    client, store = _client(monkeypatch)
    workbook = _xlsx_bytes(
        [
            ["name", "phone_number", "agent"],
            ["Alice", "+15551234567", "sales"],
        ]
    )

    response = client.post(
        "/api/outbound/campaigns/campaign-1/leads/import",
        files={
            "file": (
                "leads.xlsx",
                workbook,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200, response.text
    assert response.json()["accepted"] == 1
    campaign_id, csv_bytes, kwargs = store.imports[0]
    assert campaign_id == "campaign-1"
    assert "+15551234567" in csv_bytes.decode("utf-8")
    assert kwargs["known_agents"] == ["sales"]
    assert kwargs["known_contexts"] == ["Sales Team", "sales"]


def test_import_preserves_valid_empty_agent_set_for_validation(monkeypatch):
    client, store = _client(monkeypatch)
    monkeypatch.setattr(
        outbound_api,
        "_load_known_agent_selectors",
        lambda: ([], []),
    )

    response = client.post(
        "/api/outbound/campaigns/campaign-1/leads/import",
        files={"file": ("leads.csv", b"phone_number,agent\n2765,missing\n", "text/csv")},
    )

    assert response.status_code == 200, response.text
    _, _, kwargs = store.imports[0]
    assert kwargs["known_agents"] == []
    assert kwargs["known_contexts"] == []


def test_xlsx_conversion_rejects_invalid_archive_and_row_overflow(monkeypatch):
    with pytest.raises(ValueError, match="Invalid .xlsx workbook"):
        outbound_api._xlsx_to_csv_bytes(b"not-a-workbook")

    monkeypatch.setattr(outbound_api, "_lead_import_max_rows", lambda: 1)
    workbook = _xlsx_bytes(
        [
            ["name", "phone_number"],
            ["Alice", "+15551234567"],
            ["Bob", "+15557654321"],
        ]
    )
    with pytest.raises(ValueError, match="exceeds the 1 lead row limit"):
        outbound_api._xlsx_to_csv_bytes(workbook)


def test_manual_lead_endpoint_reuses_import_validation(monkeypatch):
    client, store = _client(monkeypatch)

    response = client.post(
        "/api/outbound/campaigns/campaign-1/leads",
        json={
            "name": "Alice",
            "phone_number": "+15551234567",
            "agent": "sales",
            "timezone": "America/Phoenix",
            "caller_id": "6789",
            "custom_vars": {"account_id": "A-1002"},
        },
    )

    assert response.status_code == 200, response.text
    _, csv_bytes, _ = store.imports[0]
    row = list(csv.DictReader(io.StringIO(csv_bytes.decode("utf-8"))))[0]
    assert row["phone_number"] == "+15551234567"
    assert row["agent"] == "sales"
    assert json.loads(row["custom_vars"]) == {"account_id": "A-1002"}


def test_lead_import_rejects_unsupported_extension_and_oversize(monkeypatch):
    client, store = _client(monkeypatch)

    unsupported = client.post(
        "/api/outbound/campaigns/campaign-1/leads/import",
        files={"file": ("leads.xls", b"legacy", "application/vnd.ms-excel")},
    )
    monkeypatch.setattr(outbound_api, "_lead_import_max_bytes", lambda: 2)
    oversized = client.post(
        "/api/outbound/campaigns/campaign-1/leads/import",
        files={"file": ("leads.csv", b"abc", "text/csv")},
    )

    assert unsupported.status_code == 400
    assert oversized.status_code == 413
    assert store.imports == []
