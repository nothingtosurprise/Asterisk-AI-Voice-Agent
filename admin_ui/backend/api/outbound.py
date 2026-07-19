"""
Outbound Campaign Dialer API endpoints (Milestone 22).

MVP scope:
- Campaign CRUD + status transitions (running/paused/stopped)
- CSV/XLSX lead import and manual lead entry (skip_existing default)
- Leads list + ignore/recycle/delete
- Attempts list + basic stats
- Voicemail drop media upload + WAV preview (for browser playback)
- Optional consent media upload + WAV preview (for browser playback)
"""

import audioop
import csv
import io
import json
import logging
import os
import re
import sys
import uuid
import wave
import zipfile
from datetime import date, datetime, time as datetime_time, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel, Field

# Add the project root before importing shared engine modules. The resolved
# fallback keeps direct Admin-backend test runs working outside containers.
project_root = os.environ.get("PROJECT_ROOT") or str(Path(__file__).resolve().parents[3])
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from src.audio.resampler import resample_audio
from src.core.outbound_schedule import normalize_outbound_daily_window

try:
    from zoneinfo import ZoneInfo, available_timezones
except Exception:  # pragma: no cover
    ZoneInfo = None  # type: ignore
    available_timezones = None  # type: ignore

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/outbound", tags=["outbound"])

def _dotenv_value(key: str) -> Optional[str]:
    """Read a key from the project's `.env` file (best-effort)."""
    try:
        env_path = os.path.join(project_root, ".env")
        if not os.path.exists(env_path):
            return None
        from dotenv import dotenv_values

        raw = dotenv_values(env_path)
        val = raw.get(key)
        if val is None:
            return None
        return str(val).strip()
    except Exception:
        return None


def _get_outbound_store():
    try:
        from src.core.outbound_store import get_outbound_store
        return get_outbound_store()
    except ImportError as e:
        logger.error("Failed to import outbound_store module: %s", e)
        raise HTTPException(status_code=500, detail="Outbound dialer module not available")


def _media_dir() -> str:
    # SECURITY: Keep media dir anchored to the known, docker-mounted location.
    # Avoid using a fully user-controlled path via env var (CodeQL path-injection).
    return "/mnt/asterisk_media/ai-generated"

def _vm_upload_max_bytes() -> int:
    try:
        # Default: 12MB (enough for ~30s stereo 44.1k WAV) while still preventing abuse.
        return max(1, int(os.getenv("AAVA_VM_UPLOAD_MAX_BYTES", "12582912")))
    except Exception:
        return 12582912


def _lead_import_max_bytes() -> int:
    try:
        return max(1, int(os.getenv("AAVA_OUTBOUND_LEAD_IMPORT_MAX_BYTES", "10485760")))
    except Exception:
        return 10485760


def _lead_import_max_rows() -> int:
    try:
        return max(1, min(100_000, int(os.getenv("AAVA_OUTBOUND_LEAD_IMPORT_MAX_ROWS", "10000"))))
    except Exception:
        return 10000


def _xlsx_cell_text(cell: Any) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, datetime_time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        number_format = str(getattr(cell, "number_format", "") or "")
        if re.fullmatch(r"0+", number_format):
            return str(value).zfill(len(number_format))
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else format(value, ".15g")
    return str(value)


def _xlsx_to_csv_bytes(data: bytes) -> bytes:
    """Convert the first XLSX worksheet into bounded UTF-8 CSV for one importer."""
    if not data:
        raise ValueError("Excel workbook is empty")
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            entries = archive.infolist()
            if len(entries) > 1000:
                raise ValueError("Excel workbook contains too many archive entries")
            if sum(int(entry.file_size or 0) for entry in entries) > 50 * 1024 * 1024:
                raise ValueError("Excel workbook expands beyond the 50 MB safety limit")
    except zipfile.BadZipFile as exc:
        raise ValueError("Invalid .xlsx workbook") from exc

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(
            io.BytesIO(data),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise ValueError(f"Unable to read .xlsx workbook: {exc}") from exc

    try:
        if not workbook.worksheets:
            raise ValueError("Excel workbook has no worksheets")
        worksheet = workbook.worksheets[0]
        max_rows = _lead_import_max_rows() + 1  # include header
        max_columns = 64
        if int(worksheet.max_row or 0) > max_rows:
            raise ValueError(f"Excel worksheet exceeds the {max_rows - 1} lead row limit")
        if int(worksheet.max_column or 0) > max_columns:
            raise ValueError(f"Excel worksheet exceeds the {max_columns} column limit")

        rows: List[List[str]] = []
        for cells in worksheet.iter_rows(
            min_row=1,
            max_row=max_rows,
            max_col=max_columns,
        ):
            row = [_xlsx_cell_text(cell) for cell in cells]
            while row and row[-1] == "":
                row.pop()
            if row or rows:
                rows.append(row)
        while rows and not any(value.strip() for value in rows[-1]):
            rows.pop()
        if not rows or not any(value.strip() for value in rows[0]):
            raise ValueError("Excel worksheet is missing a header row")

        output = io.StringIO(newline="")
        writer = csv.writer(output, lineterminator="\n")
        writer.writerows(rows)
        return output.getvalue().encode("utf-8")
    finally:
        workbook.close()


DEFAULT_CONSENT_MEDIA_URI = "sound:ai-generated/aava-consent-default"
DEFAULT_VOICEMAIL_MEDIA_URI = "sound:ai-generated/aava-voicemail-default"

class RecordingRow(BaseModel):
    media_uri: str
    filename: str
    size_bytes: int = 0


def _find_media_ulaw_path(base: str) -> Optional[str]:
    """Resolve a media basename while preserving basename case."""
    media_dir = _media_dir()
    try:
        entries = os.listdir(media_dir)
        exact_name = f"{base}.ulaw"
        exact_path = os.path.join(media_dir, exact_name)
        if exact_name in entries and os.path.isfile(exact_path):
            return exact_path

        for entry in entries:
            stem, suffix = os.path.splitext(entry)
            if stem != base or suffix.lower() != ".ulaw":
                continue
            full_path = os.path.join(media_dir, entry)
            if os.path.isfile(full_path):
                return full_path
    except FileNotFoundError:
        return None
    return None


def _media_uri_exists(media_uri: str) -> bool:
    uri = (media_uri or "").strip()
    if not uri.startswith("sound:ai-generated/"):
        return False
    base = os.path.basename(uri.split("sound:ai-generated/", 1)[1].strip())
    if not base:
        return False
    if (not _SAFE_NAME_RE.match(base)) or (".." in base):
        return False
    return _find_media_ulaw_path(base) is not None

def _safe_ai_generated_basename(media_uri: str) -> str:
    uri = (media_uri or "").strip()
    if not uri.startswith("sound:ai-generated/"):
        raise HTTPException(status_code=400, detail="media_uri must be in sound:ai-generated/")
    raw_base = uri.split("sound:ai-generated/", 1)[1].strip()
    base = os.path.basename(raw_base)
    if not base:
        raise HTTPException(status_code=400, detail="Invalid media_uri")
    # SECURITY: sanitize and reject path traversal attempts.
    if base != raw_base or (".." in base) or (not _SAFE_NAME_RE.match(base)):
        raise HTTPException(status_code=400, detail="Invalid media_uri basename")
    return base


_SAFE_NAME_RE = re.compile(r"^[a-zA-Z0-9_.-]+$")

def _ulaw_to_wav_bytes(ulaw_data: bytes) -> bytes:
    pcm16 = audioop.ulaw2lin(ulaw_data, 2)
    buf = io.BytesIO()
    with wave.open(buf, "wb") as wavf:
        wavf.setnchannels(1)
        wavf.setsampwidth(2)
        wavf.setframerate(8000)
        wavf.writeframes(pcm16)
    return buf.getvalue()

def _read_media_ulaw(media_uri: str) -> bytes:
    base = _safe_ai_generated_basename(media_uri)
    ulaw_path = _find_media_ulaw_path(base)
    if ulaw_path:
        with open(ulaw_path, "rb") as f:
            return f.read()
    raise HTTPException(status_code=404, detail="Media file not found on server")

def _convert_upload_to_ulaw(data: bytes, ext: str) -> bytes:
    if not data:
        raise HTTPException(status_code=400, detail="Empty upload")
    max_bytes = _vm_upload_max_bytes()
    if len(data) > max_bytes:
        raise HTTPException(status_code=400, detail=f"Upload too large (max {max_bytes} bytes)")

    if ext == ".ulaw":
        return data

    try:
        with wave.open(io.BytesIO(data), "rb") as wavf:
            nch = wavf.getnchannels()
            sampwidth = wavf.getsampwidth()
            fr = wavf.getframerate()
            nframes = wavf.getnframes()
            frames = wavf.readframes(nframes)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid WAV file: {e}")

    if nch not in (1, 2):
        raise HTTPException(status_code=400, detail="WAV must be mono or stereo (1–2 channels)")
    if sampwidth not in (1, 2, 3, 4):
        raise HTTPException(status_code=400, detail="Unsupported WAV sample width")

    if sampwidth != 2:
        frames = audioop.lin2lin(frames, sampwidth, 2)
    if nch == 2:
        frames = audioop.tomono(frames, 2, 0.5, 0.5)
    if fr != 8000:
        frames, _ = resample_audio(frames, fr, 8000)
    return audioop.lin2ulaw(frames, 2)

@router.get("/recordings", response_model=List[RecordingRow])
async def list_recordings():
    """
    List available recordings in the shared media directory.

    These are selectable by campaigns and can be reused across campaigns by referencing `media_uri`.
    """
    media_dir = _media_dir()
    try:
        os.makedirs(media_dir, exist_ok=True)
    except Exception:
        pass

    rows: List[RecordingRow] = []
    try:
        for entry in sorted(os.listdir(media_dir)):
            if not entry.lower().endswith(".ulaw"):
                continue
            filename = entry
            base = entry[: -len(Path(entry).suffix)] if Path(entry).suffix else entry
            path = os.path.join(media_dir, entry)
            try:
                size_bytes = int(os.path.getsize(path))
            except Exception:
                size_bytes = 0
            rows.append(
                RecordingRow(
                    media_uri=f"sound:ai-generated/{base}",
                    filename=filename,
                    size_bytes=size_bytes,
                )
            )
    except Exception:
        return []
    return rows

@router.get("/recordings/preview.wav")
async def preview_recording_wav(media_uri: str = Query(...)):
    """
    Preview any `sound:ai-generated/*` recording as WAV (browser-playable).
    Useful for Create Campaign flow (no campaign_id yet).
    """
    ulaw_data = _read_media_ulaw(media_uri)
    wav_bytes = _ulaw_to_wav_bytes(ulaw_data)
    return Response(content=wav_bytes, media_type="audio/wav")

@router.post("/recordings/upload")
async def upload_recording_to_library(kind: str = Query("generic"), file: UploadFile = File(...)):
    """
    Upload a recording to the shared library (`AAVA_MEDIA_DIR`), returning its `media_uri`.

    - Accepts `.ulaw` (8kHz μ-law) or `.wav` (PCM; auto-converted to 8kHz μ-law).
    - Enforces max upload size via `AAVA_VM_UPLOAD_MAX_BYTES`.
    """
    filename = (file.filename or "").strip() or "recording.ulaw"
    ext = os.path.splitext(filename)[1].lower().strip()
    if ext not in (".ulaw", ".wav"):
        raise HTTPException(status_code=400, detail="Upload must be .ulaw (8kHz μ-law) or .wav (PCM) audio")

    raw_name = os.path.basename(filename)
    if raw_name and not _SAFE_NAME_RE.match(raw_name):
        raise HTTPException(status_code=400, detail="Invalid filename")

    media_dir = _media_dir()
    os.makedirs(media_dir, exist_ok=True)
    unique = f"outbound-recording-{uuid.uuid4().hex[:10]}.ulaw"
    path = os.path.join(media_dir, unique)

    data = await file.read()
    ulaw_data = _convert_upload_to_ulaw(data, ext)

    with open(path, "wb") as f:
        f.write(ulaw_data)

    media_uri = f"sound:ai-generated/{unique[:-5]}"
    return {"media_uri": media_uri}

def _detect_server_timezone() -> str:
    """
    Best-effort detection of server timezone as an IANA string.
    Prefer explicit env var (TZ or AAVA_SERVER_TIMEZONE), then /etc/localtime symlink, then /etc/timezone.
    """
    def _validate_iana(tz: str) -> Optional[str]:
        tz = (tz or "").strip()
        if not tz:
            return None
        if tz.upper() == "UTC":
            return "UTC"
        if ZoneInfo is None:
            return tz
        try:
            ZoneInfo(tz)
            return tz
        except Exception:
            return None

    # Prefer configured `.env` (UI saves here), then container environment.
    env_tz = _validate_iana(_dotenv_value("TZ") or "")
    if env_tz:
        return env_tz
    # Standard Docker env var
    env_tz = _validate_iana(os.getenv("TZ") or "")
    if env_tz:
        return env_tz
    env_tz = _validate_iana(_dotenv_value("AAVA_SERVER_TIMEZONE") or "")
    if env_tz:
        return env_tz
    env_tz = _validate_iana(os.getenv("AAVA_SERVER_TIMEZONE") or "")
    if env_tz:
        return env_tz

    try:
        target = os.path.realpath("/etc/localtime")
        marker = f"{os.sep}zoneinfo{os.sep}"
        if marker in target:
            tz = target.split(marker, 1)[1].strip(os.sep)
            if tz:
                validated = _validate_iana(tz)
                if validated:
                    return validated
    except Exception:
        pass

    try:
        tz = Path("/etc/timezone").read_text(encoding="utf-8").strip()
        validated = _validate_iana(tz)
        if validated:
            return validated
    except Exception:
        pass

    return "UTC"

def _try_load_active_agents() -> Optional[List[Dict[str, Any]]]:
    """Load active Agent metadata, preserving unavailable vs. valid-empty state."""
    try:
        try:
            from agents_store import AgentsStore
        except ImportError:
            from admin_ui.backend.agents_store import AgentsStore

        with AgentsStore() as store:
            rows = store.list_all()
        return [
            {
                "slug": str(row.get("slug") or "").strip(),
                "display_name": str(row.get("display_name") or "").strip(),
                "is_default": bool(row.get("is_default")),
            }
            for row in rows
            if bool(row.get("is_active")) and str(row.get("slug") or "").strip()
        ]
    except Exception:
        logger.warning(
            "Unable to load active Agents for outbound validation", exc_info=True
        )
        return None


def _load_active_agents() -> List[Dict[str, Any]]:
    """Best-effort active Agent metadata for API/UI responses."""
    return _try_load_active_agents() or []


def _load_known_agent_selectors() -> Tuple[Optional[List[str]], Optional[List[str]]]:
    """Return canonical slugs and legacy display-name selectors for validation."""
    agents = _try_load_active_agents()
    if agents is None:
        return None, None
    slugs = [
        str(agent.get("slug") or "").strip()
        for agent in agents
        if agent.get("slug")
    ]
    legacy_names = [
        str(agent.get("display_name") or "").strip()
        for agent in agents
        if agent.get("display_name")
    ]
    # AI_CONTEXT resolution accepts display names first, then exact slugs.
    return slugs, list(dict.fromkeys([*legacy_names, *slugs]))


def _validate_campaign_schedule_for_start(campaign: Dict[str, Any]) -> None:
    """Reject invalid campaign schedule data before it can enter running state."""
    tz_name = str(campaign.get("timezone") or "").strip() or "UTC"
    if ZoneInfo is not None:
        try:
            ZoneInfo(tz_name)
        except Exception as exc:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid timezone '{tz_name}'. Use an IANA timezone like 'America/Phoenix' or 'UTC'.",
            ) from exc

    start_local = normalize_outbound_daily_window(
        campaign.get("daily_window_start_local"), "09:00"
    )
    end_local = normalize_outbound_daily_window(
        campaign.get("daily_window_end_local"), "17:00"
    )
    if start_local is None or end_local is None:
        raise HTTPException(
            status_code=400,
            detail="Daily calling window must use a valid 24-hour H:MM or HH:MM value.",
        )

    parsed_absolute: Dict[str, datetime] = {}
    for field_name, label in (
        ("run_start_at_utc", "absolute run start"),
        ("run_end_at_utc", "absolute run end"),
    ):
        raw_value = str(campaign.get(field_name) or "").strip()
        if not raw_value:
            continue
        try:
            value = datetime.fromisoformat(raw_value.replace("Z", "+00:00"))
            if value.tzinfo is None:
                value = value.replace(tzinfo=timezone.utc)
            parsed_absolute[field_name] = value.astimezone(timezone.utc)
        except Exception as exc:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid {label} timestamp '{raw_value}'.",
            ) from exc

    run_start = parsed_absolute.get("run_start_at_utc")
    run_end = parsed_absolute.get("run_end_at_utc")
    if run_start and run_end and run_start > run_end:
        raise HTTPException(status_code=400, detail="Absolute run start must be before or equal to run end.")

@router.get("/meta")
async def outbound_meta():
    """
    UI helper metadata:
    - server_timezone: what the server/container thinks is the local timezone (IANA)
    - iana_timezones: list for validation/autocomplete
    """
    tz = _detect_server_timezone()
    tzs: List[str] = []
    if available_timezones is not None:
        try:
            tzs = sorted(list(available_timezones()))
        except Exception:
            tzs = []
    return {
        "server_timezone": tz,
        "iana_timezones": tzs,
        "agents": _load_active_agents(),
        "server_now_iso": datetime.now(timezone.utc).isoformat(),
        "default_amd_options": {
            "initial_silence_ms": 2000,
            "greeting_ms": 2000,
            "after_greeting_silence_ms": 1000,
            "total_analysis_time_ms": 5000,
            "minimum_word_length_ms": 100,
            "between_words_silence_ms": 50,
            "maximum_number_of_words": 10,
        },
    }


class CampaignCreateRequest(BaseModel):
    name: str = Field(..., min_length=1)
    timezone: str = "UTC"
    run_start_at_utc: Optional[str] = None
    run_end_at_utc: Optional[str] = None
    daily_window_start_local: str = "09:00"
    daily_window_end_local: str = "17:00"
    max_concurrent: int = Field(1, ge=1, le=5)
    min_interval_seconds_between_calls: int = Field(5, ge=0, le=3600)
    default_context: str = "default"
    voicemail_drop_enabled: bool = True
    voicemail_drop_mode: str = "upload"  # upload|tts
    voicemail_drop_text: Optional[str] = None
    voicemail_drop_media_uri: Optional[str] = None
    consent_enabled: bool = False
    consent_media_uri: Optional[str] = None
    consent_timeout_seconds: int = Field(5, ge=1, le=30)
    amd_options: Dict[str, Any] = Field(default_factory=dict)


class CampaignStatusRequest(BaseModel):
    status: str  # running|paused|stopped|draft|archived|completed
    cancel_pending: bool = False

class LeadRecycleRequest(BaseModel):
    mode: str = Field("redial", pattern="^(redial|reset)$")  # redial|reset


class LeadImportResponse(BaseModel):
    accepted: int = 0
    rejected: int = 0
    duplicates: int = 0
    errors: List[Dict[str, Any]] = Field(default_factory=list)
    error_csv: str = ""
    error_csv_truncated: bool = False
    warnings: List[Dict[str, Any]] = Field(default_factory=list)
    warnings_truncated: bool = False


class ManualLeadCreateRequest(BaseModel):
    phone_number: str = Field(..., min_length=1, max_length=64)
    name: Optional[str] = Field(None, max_length=200)
    agent: Optional[str] = Field(None, max_length=64)
    timezone: Optional[str] = Field(None, max_length=100)
    caller_id: Optional[str] = Field(None, max_length=64)
    custom_vars: Dict[str, Any] = Field(default_factory=dict)


@router.get("/sample.csv")
async def download_sample_csv():
    """
    Download a sample CSV for lead import.

    Columns supported by the importer (full format):
      - name (optional)
      - phone_number (required)
        - Can be E.164 (+15551234567) or an internal extension (e.g., 2765)
      - agent (optional; active Agent slug used with AI_AGENT)
      - context (deprecated compatibility alias for agent)
      - timezone (optional)
      - caller_id (optional)
      - custom_vars (optional JSON object)
    """
    active_agents = _load_active_agents()
    sample_agent = next(
        (str(agent["slug"]) for agent in active_agents if agent.get("is_default")),
        str(active_agents[0]["slug"]) if active_agents else "default",
    )
    csv_text = (
        "name,phone_number,agent,timezone,caller_id,custom_vars\n"
        f"Extension Test,2765,{sample_agent},America/Phoenix,6789,\"{{\"\"name\"\":\"\"Extension Test\"\",\"\"note\"\":\"\"Call internal extension\"\"}}\"\n"
        f"Alice Example,+15557654321,{sample_agent},America/Phoenix,6789,\"{{\"\"name\"\":\"\"Alice Example\"\",\"\"account_id\"\":\"\"A-1002\"\",\"\"note\"\":\"\"US lead example\"\"}}\"\n"
        f"International Example,+447700900123,{sample_agent},America/Phoenix,6789,\"{{\"\"name\"\":\"\"International Example\"\",\"\"account_id\"\":\"\"A-1003\"\",\"\"note\"\":\"\"International lead example\"\"}}\"\n"
    )
    return Response(
        content=csv_text,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="outbound_sample_leads.csv"'},
    )


@router.get("/campaigns")
async def list_campaigns(include_archived: bool = Query(False)):
    store = _get_outbound_store()
    return await store.list_campaigns(include_archived=bool(include_archived))


@router.post("/campaigns")
async def create_campaign(req: CampaignCreateRequest):
    store = _get_outbound_store()
    payload = req.model_dump()
    try:
        if payload.get("voicemail_drop_enabled") and not (payload.get("voicemail_drop_media_uri") or "").strip():
            if _media_uri_exists(DEFAULT_VOICEMAIL_MEDIA_URI):
                payload["voicemail_drop_media_uri"] = DEFAULT_VOICEMAIL_MEDIA_URI
        if payload.get("consent_enabled") and not (payload.get("consent_media_uri") or "").strip():
            if _media_uri_exists(DEFAULT_CONSENT_MEDIA_URI):
                payload["consent_media_uri"] = DEFAULT_CONSENT_MEDIA_URI
    except Exception:
        pass
    return await store.create_campaign(payload)


@router.get("/campaigns/{campaign_id}")
async def get_campaign(campaign_id: str):
    store = _get_outbound_store()
    try:
        return await store.get_campaign(campaign_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="Campaign not found")


@router.patch("/campaigns/{campaign_id}")
async def update_campaign(campaign_id: str, body: Dict[str, Any]):
    store = _get_outbound_store()
    try:
        return await store.update_campaign(campaign_id, body or {})
    except KeyError:
        raise HTTPException(status_code=404, detail="Campaign not found")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/campaigns/{campaign_id}/clone")
async def clone_campaign(campaign_id: str):
    store = _get_outbound_store()
    try:
        return await store.clone_campaign(campaign_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="Campaign not found")

@router.post("/campaigns/{campaign_id}/archive")
async def archive_campaign(campaign_id: str):
    store = _get_outbound_store()
    try:
        campaign = await store.get_campaign(campaign_id)
        if str(campaign.get("status") or "").strip().lower() == "running":
            raise HTTPException(status_code=400, detail="Pause/stop the campaign before archiving")
        return await store.set_campaign_status(campaign_id, "archived", cancel_pending=False)
    except KeyError:
        raise HTTPException(status_code=404, detail="Campaign not found")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/campaigns/{campaign_id}")
async def delete_campaign(campaign_id: str):
    store = _get_outbound_store()
    try:
        await store.delete_campaign(campaign_id)
        return {"ok": True}
    except KeyError:
        raise HTTPException(status_code=404, detail="Campaign not found")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/campaigns/{campaign_id}/status")
async def set_campaign_status(campaign_id: str, req: CampaignStatusRequest):
    store = _get_outbound_store()
    try:
        # Guardrails: require enabled recordings before running.
        if req.status.strip().lower() == "running":
            campaign = await store.get_campaign(campaign_id)
            stats = await store.campaign_stats(campaign_id)
            lead_states = (stats or {}).get("lead_states") or {}
            pending = int(lead_states.get("pending") or 0)
            if pending <= 0:
                canceled = int(lead_states.get("canceled") or 0)
                completed = int(lead_states.get("completed") or 0)
                raise HTTPException(
                    status_code=400,
                    detail=f"No pending leads to dial (canceled={canceled}, completed={completed}). Recycle leads back to pending, then Start again.",
                )
            if bool(int(campaign.get("voicemail_drop_enabled") or 1)):
                media_uri = (campaign.get("voicemail_drop_media_uri") or "").strip()
                if not media_uri:
                    raise HTTPException(
                        status_code=400,
                        detail="Voicemail drop is enabled but no voicemail recording is set. Upload/generate voicemail before starting.",
                    )
            if bool(int(campaign.get("consent_enabled") or 0)):
                consent_uri = (campaign.get("consent_media_uri") or "").strip()
                if not consent_uri:
                    raise HTTPException(
                        status_code=400,
                        detail="Consent gate is enabled but no consent recording is set. Upload consent before starting.",
                    )
            _validate_campaign_schedule_for_start(campaign)
        return await store.set_campaign_status(campaign_id, req.status, cancel_pending=bool(req.cancel_pending))
    except KeyError:
        raise HTTPException(status_code=404, detail="Campaign not found")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/campaigns/{campaign_id}/stats")
async def campaign_stats(campaign_id: str):
    store = _get_outbound_store()
    return await store.campaign_stats(campaign_id)


@router.post("/campaigns/{campaign_id}/leads/import", response_model=LeadImportResponse)
async def import_leads(
    campaign_id: str,
    file: UploadFile = File(...),
    skip_existing: bool = Query(True),
    max_error_rows: int = Query(20, ge=1, le=200),
):
    store = _get_outbound_store()
    try:
        max_bytes = _lead_import_max_bytes()
        data = await file.read(max_bytes + 1)
        if len(data) > max_bytes:
            raise HTTPException(
                status_code=413,
                detail=f"Lead import is too large (max {max_bytes} bytes)",
            )
        filename = os.path.basename((file.filename or "").strip()).lower()
        if filename.endswith(".xlsx"):
            data = _xlsx_to_csv_bytes(data)
        elif not filename.endswith(".csv"):
            raise HTTPException(
                status_code=400,
                detail="Lead import must be a .csv or .xlsx file",
            )
        known_agents, known_contexts = _load_known_agent_selectors()
        result = await store.import_leads_csv(
            campaign_id,
            data,
            skip_existing=bool(skip_existing),
            max_error_rows=int(max_error_rows),
            known_agents=known_agents,
            known_contexts=known_contexts,
        )
        return LeadImportResponse(**result)
    except KeyError:
        raise HTTPException(status_code=404, detail="Campaign not found")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post(
    "/campaigns/{campaign_id}/leads",
    response_model=LeadImportResponse,
)
async def add_manual_lead(campaign_id: str, req: ManualLeadCreateRequest):
    """Add one lead through the same validation and duplicate path as file imports."""
    output = io.StringIO(newline="")
    writer = csv.DictWriter(
        output,
        fieldnames=[
            "name",
            "phone_number",
            "agent",
            "timezone",
            "caller_id",
            "custom_vars",
        ],
        lineterminator="\n",
    )
    writer.writeheader()
    writer.writerow(
        {
            "name": (req.name or "").strip(),
            "phone_number": req.phone_number.strip(),
            "agent": (req.agent or "").strip(),
            "timezone": (req.timezone or "").strip(),
            "caller_id": (req.caller_id or "").strip(),
            "custom_vars": json.dumps(req.custom_vars or {}, separators=(",", ":")),
        }
    )

    store = _get_outbound_store()
    try:
        known_agents, known_contexts = _load_known_agent_selectors()
        result = await store.import_leads_csv(
            campaign_id,
            output.getvalue().encode("utf-8"),
            skip_existing=True,
            max_error_rows=20,
            known_agents=known_agents,
            known_contexts=known_contexts,
        )
        return LeadImportResponse(**result)
    except KeyError:
        raise HTTPException(status_code=404, detail="Campaign not found")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/campaigns/{campaign_id}/leads")
async def list_leads(
    campaign_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    state: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
):
    store = _get_outbound_store()
    return await store.list_leads(campaign_id, page=page, page_size=page_size, state=state, q=q)


@router.post("/leads/{lead_id}/cancel")
async def cancel_lead(lead_id: str):
    store = _get_outbound_store()
    ok = await store.cancel_lead(lead_id)
    if not ok:
        raise HTTPException(status_code=400, detail="Lead cannot be canceled in its current state")
    return {"ok": True}

@router.post("/leads/{lead_id}/ignore")
async def ignore_lead(lead_id: str):
    store = _get_outbound_store()
    ok = await store.ignore_lead(lead_id)
    if not ok:
        raise HTTPException(status_code=400, detail="Lead cannot be ignored in its current state")
    return {"ok": True}

@router.post("/leads/{lead_id}/recycle")
async def recycle_lead(lead_id: str, req: LeadRecycleRequest):
    store = _get_outbound_store()
    ok = await store.recycle_lead(lead_id, mode=req.mode)
    if not ok:
        raise HTTPException(status_code=400, detail="Lead cannot be recycled in its current state")
    return {"ok": True}

@router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str):
    store = _get_outbound_store()
    try:
        await store.delete_lead(lead_id)
        return {"ok": True}
    except KeyError:
        raise HTTPException(status_code=404, detail="Lead not found")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/campaigns/{campaign_id}/attempts")
async def list_attempts(
    campaign_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
):
    store = _get_outbound_store()
    return await store.list_attempts(campaign_id, page=page, page_size=page_size)


@router.post("/campaigns/{campaign_id}/voicemail/upload")
async def upload_voicemail_media(campaign_id: str, file: UploadFile = File(...)):
    store = _get_outbound_store()
    filename = (file.filename or "").strip() or "voicemail.ulaw"
    ext = os.path.splitext(filename)[1].lower().strip()
    if ext not in (".ulaw", ".wav"):
        raise HTTPException(status_code=400, detail="Upload must be .ulaw (8kHz μ-law) or .wav (PCM) audio")

    raw_name = os.path.basename(filename)
    if not _SAFE_NAME_RE.match(raw_name):
        raise HTTPException(status_code=400, detail="Invalid filename")

    media_dir = _media_dir()
    os.makedirs(media_dir, exist_ok=True)
    unique = f"outbound-vm-{uuid.uuid4().hex[:10]}.ulaw"
    path = os.path.join(media_dir, unique)
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty upload")
    max_bytes = _vm_upload_max_bytes()
    if len(data) > max_bytes:
        raise HTTPException(status_code=400, detail=f"Upload too large (max {max_bytes} bytes)")

    if ext == ".ulaw":
        ulaw_data = data
    else:
        # Convert WAV (PCM) -> 8kHz μ-law so Asterisk Playback() can use it directly.
        try:
            with wave.open(io.BytesIO(data), "rb") as wavf:
                nch = wavf.getnchannels()
                sampwidth = wavf.getsampwidth()
                fr = wavf.getframerate()
                nframes = wavf.getnframes()
                frames = wavf.readframes(nframes)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid WAV file: {e}")

        if nch not in (1, 2):
            raise HTTPException(status_code=400, detail="WAV must be mono or stereo (1–2 channels)")
        if sampwidth not in (1, 2, 3, 4):
            raise HTTPException(status_code=400, detail="Unsupported WAV sample width")

        # Normalize to 16-bit little-endian PCM for processing.
        if sampwidth != 2:
            frames = audioop.lin2lin(frames, sampwidth, 2)
        if nch == 2:
            # Downmix stereo -> mono.
            frames = audioop.tomono(frames, 2, 0.5, 0.5)

        # Resample to 8kHz if needed.
        if fr != 8000:
            frames, _ = resample_audio(frames, fr, 8000)

        ulaw_data = audioop.lin2ulaw(frames, 2)

    with open(path, "wb") as f:
        f.write(ulaw_data)

    media_uri = f"sound:ai-generated/{unique[:-5]}"
    try:
        campaign = await store.update_campaign(campaign_id, {"voicemail_drop_media_uri": media_uri})
    except KeyError:
        raise HTTPException(status_code=404, detail="Campaign not found")
    return {"media_uri": media_uri, "campaign": campaign}

@router.post("/campaigns/{campaign_id}/consent/upload")
async def upload_consent_media(campaign_id: str, file: UploadFile = File(...)):
    store = _get_outbound_store()
    filename = (file.filename or "").strip() or "consent.ulaw"
    ext = os.path.splitext(filename)[1].lower().strip()
    if ext not in (".ulaw", ".wav"):
        raise HTTPException(status_code=400, detail="Upload must be .ulaw (8kHz μ-law) or .wav (PCM) audio")

    raw_name = os.path.basename(filename)
    if not _SAFE_NAME_RE.match(raw_name):
        raise HTTPException(status_code=400, detail="Invalid filename")

    media_dir = _media_dir()
    os.makedirs(media_dir, exist_ok=True)
    unique = f"outbound-consent-{uuid.uuid4().hex[:10]}.ulaw"
    path = os.path.join(media_dir, unique)
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty upload")
    max_bytes = _vm_upload_max_bytes()
    if len(data) > max_bytes:
        raise HTTPException(status_code=400, detail=f"Upload too large (max {max_bytes} bytes)")

    if ext == ".ulaw":
        ulaw_data = data
    else:
        try:
            with wave.open(io.BytesIO(data), "rb") as wavf:
                nch = wavf.getnchannels()
                sampwidth = wavf.getsampwidth()
                fr = wavf.getframerate()
                nframes = wavf.getnframes()
                frames = wavf.readframes(nframes)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid WAV file: {e}")

        if nch not in (1, 2):
            raise HTTPException(status_code=400, detail="WAV must be mono or stereo (1–2 channels)")
        if sampwidth not in (1, 2, 3, 4):
            raise HTTPException(status_code=400, detail="Unsupported WAV sample width")

        if sampwidth != 2:
            frames = audioop.lin2lin(frames, sampwidth, 2)
        if nch == 2:
            frames = audioop.tomono(frames, 2, 0.5, 0.5)
        if fr != 8000:
            frames, _ = resample_audio(frames, fr, 8000)
        ulaw_data = audioop.lin2ulaw(frames, 2)

    with open(path, "wb") as f:
        f.write(ulaw_data)

    media_uri = f"sound:ai-generated/{unique[:-5]}"
    try:
        campaign = await store.update_campaign(campaign_id, {"consent_media_uri": media_uri})
    except KeyError:
        raise HTTPException(status_code=404, detail="Campaign not found")
    return {"media_uri": media_uri, "campaign": campaign}


@router.get("/campaigns/{campaign_id}/voicemail/preview.wav")
async def preview_voicemail_wav(campaign_id: str):
    store = _get_outbound_store()
    try:
        campaign = await store.get_campaign(campaign_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="Campaign not found")

    media_uri = (campaign.get("voicemail_drop_media_uri") or "").strip()
    ulaw_data = _read_media_ulaw(media_uri)

    wav_bytes = _ulaw_to_wav_bytes(ulaw_data)
    return Response(content=wav_bytes, media_type="audio/wav")


@router.get("/campaigns/{campaign_id}/consent/preview.wav")
async def preview_consent_wav(campaign_id: str):
    store = _get_outbound_store()
    try:
        campaign = await store.get_campaign(campaign_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="Campaign not found")

    media_uri = (campaign.get("consent_media_uri") or "").strip()
    ulaw_data = _read_media_ulaw(media_uri)

    wav_bytes = _ulaw_to_wav_bytes(ulaw_data)
    return Response(content=wav_bytes, media_type="audio/wav")
